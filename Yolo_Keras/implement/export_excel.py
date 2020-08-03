import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time
from datetime import datetime

# def save(no_obj_list = [],not_match_list = [], n = 0):
#     wb = openpyxl.load_workbook('export.xlsx')
#     fontObj1 = Font(name='Times New Roman', size=20,bold=True)
#     title = 'ver' + str(n)
#     wb.create_sheet(index=0,title = title)
#     now = datetime.now()
#     dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
#     sheet = wb.active
#     sheet.merge_cells('C1:H1')
#     sheet['C1'] = 'PRODUCTS MANAGEMENT'
#     sheet['C1'].font = fontObj1
#     sheet['C1'].alignment = Alignment(horizontal='center')
#     sheet.merge_cells('B2:D2')
#     sheet.column_dimensions['E'].width = 20
#     sheet.column_dimensions['F'].width = 20
#     #
#     #
#     sheet['A2'] = 'Time'
#     sheet['B2'] = dt_string
#
#     # for i in range(len(no_obj_list)):
#     #     sheet['A'+str(i+3)] = no_obj_list[i]
#     #     sheet['B' + str(i + 3)] = 'no_obj'
#     #
#     #
#     # for idx,i in enumerate(range(len(no_obj_list),len(no_obj_list)+len(not_match_list))):
#     #     sheet['A' + str(i + 3)] = not_match_list[idx]
#     #     sheet['B' + str(i + 3)] = 'not_match'
#
#     for i in range(len(no_obj_list)):
#         sheet['E3'] = 'No_Object'
#         sheet['E' + str(i + 4)] = no_obj_list[i]
#
#
#     for i in range(len(not_match_list)):
#         sheet['F' + str(i + 4)] = not_match_list[i]
#         sheet['F3'] = 'Not_match'
#
#     length = max(len(no_obj_list),len(not_match_list))+3
#     mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
#                                                           showRowStripes=False)
#     # create a table
#     table = openpyxl.worksheet.table.Table(ref='E3'+':F'+str(length),
#                                            displayName='No_ObjectNot_Match',
#                                            tableStyleInfo=mediumStyle)
#     # add the table to the worksheet
#     sheet.add_table(table)
#
#     file_name = 'export.xlsx'
#     wb.save(file_name)

def save(no_obj_list = [],not_match_list = [],_no_object_barcode_save = [],_not_match_barcode_save = [],n = 0):
    print(_no_object_barcode_save)
    wb = openpyxl.load_workbook('export.xlsx')
    fontObj1 = Font(name='Times New Roman', size=20,bold=True)
    title = 'ver' + str(n)
    wb.create_sheet(index=0,title = title)
    now = datetime.now()
    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    sheet = wb.active
    sheet.merge_cells('C1:H1')
    sheet['C1'] = 'PRODUCTS MANAGEMENT'
    sheet['C1'].font = fontObj1
    sheet['C1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('B2:D2')
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    #
    #
    sheet['A2'] = 'Time'
    sheet['B2'] = dt_string

    # for i in range(len(no_obj_list)):
    #     sheet['A'+str(i+3)] = no_obj_list[i]
    #     sheet['B' + str(i + 3)] = 'no_obj'
    #
    #
    # for idx,i in enumerate(range(len(no_obj_list),len(no_obj_list)+len(not_match_list))):
    #     sheet['A' + str(i + 3)] = not_match_list[idx]
    #     sheet['B' + str(i + 3)] = 'not_match'

    for i in range(len(no_obj_list)):
        sheet['E3'] = 'No_Object'

        sheet['E' + str(i + 4)] = no_obj_list[i]

    for i in range(len(_no_object_barcode_save)):

        sheet['F' + str(i + 4)] = str(_no_object_barcode_save[i])
        sheet['F3'] = 'Right_Infor'

    for i in range(len(not_match_list)):
        sheet['G' + str(i + 4)] = not_match_list[i]
        sheet['G3'] = 'Not_match'
        sheet['H3'] = 'Right_Infor'
        sheet['H'+str(i+4)] = str(_not_match_barcode_save[i])


    length = max(len(no_obj_list),len(not_match_list))+3
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                          showRowStripes=False)
    # create a table
    table = openpyxl.worksheet.table.Table(ref='E3'+':H'+str(length),
                                           displayName='No_ObjectRight_InforNot_MatchRight_Infor',
                                           tableStyleInfo=mediumStyle)
    # add the table to the worksheet
    sheet.add_table(table)

    file_name = 'export.xlsx'
    wb.save(file_name)

